# -*- coding: utf-8 -*-
"""Xuat ma tran nhan qua (core/ce_matrix.py) ra file .xlsx dinh dang TAI LIEU:
dong bang tieu de, to mau o OR/AND, va 1 sheet phu ghi nguon du lieu (danh sach DB,
ngay dung) de nop duoc."""
from __future__ import annotations
import os
import datetime

from . import io_point as IOP

# Cot co dinh dung TRUOC cac cot tin hieu dich. Cua so ma tran (ui/ce_matrix_dialog)
# dung chung danh sach nay - de them/bot cot khong lam lech chi so o danh dau.
COT_CO_DINH = ["Nguyen nhan goc", "KKS", "Diem vao/ra", "Nguon", "CPU"]
NCD = len(COT_CO_DINH)


def export_matrix(path, columns, rows, db_paths=None):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CE Matrix"

    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    or_font = Font(color="1D4ED8", bold=True, size=12)
    and_font = Font(color="B45309", bold=True, size=11)
    or_fill = PatternFill("solid", fgColor="DBEAFE")
    and_fill = PatternFill("solid", fgColor="FEF3C7")
    raw_font = Font(color="78716C", italic=True)
    zebra = PatternFill("solid", fgColor="F8FAFC")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(COT_CO_DINH) + list(columns)
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = box
    ws.freeze_panes = "B2"          # giu ca hang tieu de VA cot ten nguyen nhan
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), max(1, len(rows) + 1))
    ws.row_dimensions[1].height = 30

    n_or = n_and = 0
    for ri, row in enumerate(rows, start=2):
        c1 = ws.cell(row=ri, column=1, value=row.get("label"))
        c1.alignment = wrap
        if row.get("raw_name"):
            c1.font = raw_font
        # Nguong la 1 DUONG CONG F(x) chu khong phai 1 con so: dinh ca cum diem gay
        # khuc vao o de nguoi doc tai lieu tra tai cho, khong phai mo lai ban ve.
        # Kem theo la ly lich diem do hien truong (tag I/O, dia chi, he thong dau kia).
        ghi = [x for x in [row.get("note"), IOP.chu_thich(row.get("io"))] if x]
        if ghi:
            c1.comment = Comment("\n\n".join(ghi), "T-Designer", width=460, height=230)
        ws.cell(row=ri, column=2, value=IOP.ma_kks(row.get("io")))
        ws.cell(row=ri, column=3, value=IOP.mo_ta(row.get("io")))
        ws.cell(row=ri, column=4,
                value=("Tai lieu (TAG)" if row.get("source") == "tag" else "Suy luan tu day"))
        ws.cell(row=ri, column=5, value=row.get("source_txt")
                or (row.get("sheetlbl") or row.get("sheet") or ""))
        for ci, col in enumerate(columns, start=NCD + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = center
            cell.border = box
            m = row.get("marks", {}).get(col)
            if not m:
                if ri % 2 == 0:
                    cell.fill = zebra
                continue
            if m["kind"] == "or":
                cell.value = "●"
                cell.font = or_font
                cell.fill = or_fill
                cell.comment = Comment("Mot minh tin hieu nay da du gay ra '%s'" % col,
                                       "T-Designer")
                n_or += 1
            else:
                cell.value = "▲%s" % m.get("group", "")
                cell.font = and_font
                cell.fill = and_fill
                with_txt = "\n  + ".join(m.get("with") or [])
                cell.comment = Comment(
                    ("Chi gay ra '%s' khi KET HOP du:\n  + %s" % (col, with_txt))
                    if with_txt else "Thuoc 1 nhom AND", "T-Designer")
                n_and += 1
        for ci in range(1, NCD + 1):
            ws.cell(row=ri, column=ci).border = box

    for ci, rong in zip(range(1, NCD + 1), (52, 20, 26, 16, 30)):
        ws.column_dimensions[get_column_letter(ci)].width = rong
    for ci in range(NCD + 1, NCD + 1 + len(columns)):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    # ---- sheet phu: nguon du lieu, de nguoi doc biet bang nay tu dau ma co ----
    ws2 = wb.create_sheet("Nguon du lieu")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 90
    bold = Font(bold=True)

    def _kv(r, k, v):
        a = ws2.cell(row=r, column=1, value=k); a.font = bold
        ws2.cell(row=r, column=2, value=v)

    r = 1
    _kv(r, "Cong cu", "T-Designer Lite - Ma tran nhan qua (Cause & Effect Matrix)"); r += 1
    _kv(r, "Ngay dung bang", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")); r += 1
    _kv(r, "Tin hieu dich", ", ".join(columns)); r += 1
    _kv(r, "So nguyen nhan", str(len(rows))); r += 1
    _kv(r, "So o OR / AND", "%d / %d" % (n_or, n_and)); r += 1
    n_io = sum(1 for x in rows if x.get("io"))
    _kv(r, "Tra ra diem hien truong", "%d / %d nguyen nhan" % (n_io, len(rows))); r += 2
    a = ws2.cell(row=r, column=1, value="File DB da dung"); a.font = bold; r += 1
    for p in (db_paths or []):
        ws2.cell(row=r, column=1, value=os.path.basename(p))
        ws2.cell(row=r, column=2, value=p)
        r += 1
    r += 1
    a = ws2.cell(row=r, column=1, value="Ghi chu"); a.font = bold
    ws2.cell(row=r, column=2,
             value="● = nguyen nhan doc lap, mot minh du gay hieu ung.  "
                   "▲Gxx = phai ket hop du ca nhom Gxx (xem chu thich trong o).  "
                   "Cot 'CPU': noi tim thay nguyen nhan.  Cot 'Nguon': 'Tai lieu (TAG)' = nhan do ky su ghi san trong CAD_TAG_FID; "
                   "'Suy luan tu day' = app tu lan nguoc theo day va khoi logic.  "
                   "Cot 'KKS' va 'Diem vao/ra': ma va ly lich diem do hien truong, "
                   "doc tu khoi dau cuoi I/O trong DB (loai DI/AI/DO/AO, dai do, kieu "
                   "tin hieu, he thong dau kia); de trong khi ten do khong ra thang 1 "
                   "diem I/O nao.")
    ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[r].height = 78

    wb.save(path)
